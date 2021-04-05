'''
References:
RPA for Python: https://github.com/tebelorg/RPA-Python
RPA Challenge: www.rpachallenge.com (move the file challenge.xlsx to the project directory) 
Openpyxl: https://openpyxl.readthedocs.io/en/stable/
XPath in Chrome: https://stackoverflow.com/questions/3030487/is-there-a-way-to-get-the-xpath-in-google-chrome

Setting up the environment:
mkdir RPA-Challenge
virtual environment: virtualenv -p python3 rpa_excel
activate: source rpa_excel/bin/activate
pip3 install rpa
pip3 install openpyxl
output all requirements: pip3 freeze > requirements.txt
'''

from openpyxl import load_workbook
import rpa as bot

# load the workbook and specify sheet
wb = load_workbook(filename='challenge.xlsx')
sheet = wb['Sheet1']

# get workbook range information
print('Maximum row/s:' + str(sheet.max_row))
print('Maximum col/s:' + str(sheet.max_column))
# NOTE1: the workfile will have more rows and columns than can be seen. it's possible 
#  file was recycled. manually delete those 'hidden' rows and columns
# NOTE2: when using lubuntu, LibreOffice Calc will not allow you to resave the file unless 
#  reformatted to .ods. Try editing the file using Gnumeric instead to retain format as .xlsx.

# loop through the rows, skip first row as the header, and print
for r in range(2, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        cell_obj = sheet.cell(row=r, column=c)
        print(cell_obj.value, end=',')
    print()

# input the values to the rpachallenge website through web automation
# this makes use of xpath (using chrome, inspect the element > right-click on highlighted element > 
# copy > copy xpath)
bot.init()
bot.url('http://www.rpachallenge.com')
# start the clock
bot.click('/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/button')
# loop through the records and input onto website
for r in range(2, sheet.max_row + 1):
    # first name
    bot.type('//*[@ng-reflect-name="labelFirstName"]', sheet.cell(row=r, column=1).value)
    # last name
    bot.type('//*[@ng-reflect-name="labelLastName"]', sheet.cell(row=r, column=2).value)
    # company name
    bot.type('//*[@ng-reflect-name="labelCompanyName"]', sheet.cell(row=r, column=3).value)
    # role in company
    bot.type('//*[@ng-reflect-name="labelRole"]', sheet.cell(row=r, column=4).value)
    # address
    bot.type('//*[@ng-reflect-name="labelAddress"]', sheet.cell(row=r, column=5).value)
    # email
    bot.type('//*[@ng-reflect-name="labelEmail"]', sheet.cell(row=r, column=6).value)
    # phone number
    bot.type('//*[@ng-reflect-name="labelPhone"]', str(sheet.cell(row=r, column=7).value))
    # click submit
    bot.click('/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input')
# get result
print('RPA Challenge result: ' + bot.read('/html/body/app-root/div[2]/app-rpa1/div/div[2]/div[2]'))
# take screenshot
bot.snap('page', 'rpa-challenge-result.png')
# close the browser
bot.close()